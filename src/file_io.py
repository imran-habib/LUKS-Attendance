"""file_io.py - Read attendance .xls/.xlsx, handle merged cells, return clean grid."""
import os
import re
from openpyxl import load_workbook


def read_attendance(filepath):
    """Read .xls or .xlsx attendance file, return structured data.
    
    Returns dict: {
        "duration": str,        # e.g. "2026/01/23 ~ 01/29"
        "days": [(col_idx, day_num, day_name), ...],  # day columns
        "employees": [{"no": str, "name": str, "punches": {day_num: str, ...}}, ...]
    }
    """
    ext = os.path.splitext(filepath)[1].lower()
    
    if ext == ".xls":
        filepath = _convert_xls_to_xlsx(filepath)
    
    wb = load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    
    # Unmerge and fill merged cell values
    _fill_merged(ws)
    
    # Row 2: duration string (col 3 in merged range C2:AA2)
    duration = str(ws.cell(2, 3).value or "")
    
    # Row 3: day numbers (cols 3+)
    # Row 4: day names (Fr, Sa, Su, Mo, Tu, We, Th)
    days = []
    for c in range(3, ws.max_column + 1):
        v = ws.cell(3, c).value
        if v and str(v).strip().replace(".0", "").isdigit():
            day_num = int(float(str(v).strip()))
            day_name = str(ws.cell(4, c).value or "").strip()
            days.append((c, day_num, day_name))
    
    # Rows 5+: employee data
    employees = []
    for r in range(5, ws.max_row + 1):
        name = str(ws.cell(r, 2).value or "").strip()
        if not name:
            continue
        emp_no = str(ws.cell(r, 1).value or "").strip().replace(".0", "")
        punches = {}
        for col, day_num, _ in days:
            cell_val = ws.cell(r, col).value
            if cell_val:
                punches[day_num] = str(cell_val).strip()
        employees.append({"no": emp_no, "name": name, "punches": punches})
    
    wb.close()
    return {"duration": duration, "days": days, "employees": employees}


def _fill_merged(ws):
    """Fill all cells in merged ranges with the top-left value, then unmerge."""
    for mr in list(ws.merged_cells.ranges):
        val = ws.cell(mr.min_row, mr.min_col).value
        ws.unmerge_cells(str(mr))
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                ws.cell(r, c).value = val


def _convert_xls_to_xlsx(filepath):
    """Convert .xls to .xlsx using xlrd + openpyxl, return new path."""
    import xlrd
    from openpyxl import Workbook
    
    xls = xlrd.open_workbook(filepath)
    wb = Workbook()
    
    for i, sheet in enumerate(xls.sheets()):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = sheet.name
        for r in range(sheet.nrows):
            for c in range(sheet.ncols):
                ws.cell(r + 1, c + 1).value = sheet.cell_value(r, c)
    
    out_path = filepath + "x"  # .xls -> .xlsx
    wb.save(out_path)
    wb.close()
    return out_path
