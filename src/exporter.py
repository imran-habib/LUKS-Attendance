"""exporter.py - Generate final .xlsx with Attendance, Salary, and Employee DB tabs."""
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter


THIN_BORDER = Border(
    left=Side("thin"), right=Side("thin"),
    top=Side("thin"), bottom=Side("thin")
)
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
RED_FILL = PatternFill("solid", fgColor="FF6666")
BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center")


def generate_output(attendance_rows, salary_records, employee_db, output_path,
                    flagged_rows=None):
    """Write final .xlsx with 3 sheets: Attendance, Salary, Employee DB.
    
    attendance_rows: [{name, day, in_time, out_time, worked, ot, deduction, status}, ...]
    salary_records: [{name, days, ot_hours, ded_hours, net_hours, daily_rate, hourly_rate, advance, arrears, net_salary}, ...]
    employee_db: {name_lower: {"daily_rate": int, "type": "weekly"|"monthly"|"excluded"}, ...}
    flagged_rows: set of (name, day) that need HR verification
    """
    wb = Workbook()
    flagged_rows = flagged_rows or set()
    
    # --- Attendance Sheet ---
    ws_att = wb.active
    ws_att.title = "Attendance"
    headers_att = ["Name", "Day", "IN", "OUT", "Worked", "OT", "Deduction", "Status"]
    ws_att.append(headers_att)
    
    for row in attendance_rows:
        ws_att.append([
            row["name"], row["day"], row.get("in_time", ""),
            row.get("out_time", ""), row.get("worked", ""),
            row.get("ot", ""), row.get("deduction", ""),
            row.get("status", "")
        ])
    
    # Highlight flagged rows
    for r in range(2, ws_att.max_row + 1):
        name = ws_att.cell(r, 1).value
        day = ws_att.cell(r, 2).value
        status = ws_att.cell(r, 8).value
        if (name, day) in flagged_rows or status in ("multi_punch_verify", "hr_entered"):
            fill = YELLOW_FILL if status == "multi_punch_verify" else RED_FILL
            for c in range(1, 9):
                ws_att.cell(r, c).fill = fill
    
    _style_sheet(ws_att)
    
    # --- Salary Sheet ---
    ws_sal = wb.create_sheet("Salary")
    headers_sal = ["Name", "Days", "OT (hrs)", "Deduction (hrs)", "Net Hours",
                   "Daily Rate", "Hourly Rate", "Extra Hrs", "Advance", "Arrears", "Net Salary"]
    ws_sal.append(headers_sal)
    
    for i, rec in enumerate(salary_records, 2):
        ws_sal.append([
            rec["name"], rec["days"], rec["ot_hours"], rec["ded_hours"],
            rec["net_hours"], rec["daily_rate"], rec["hourly_rate"],
            0,  # Extra Hrs (manual HR field)
            rec["advance"], rec["arrears"], None  # formula
        ])
        # Net Salary formula: Days*DailyRate + (NetHours+ExtraHrs)*HourlyRate - Advance + Arrears
        ws_sal[f"K{i}"] = f"=ROUND(B{i}*F{i}+(E{i}+H{i})*G{i}-I{i}+J{i},2)"
    
    # Total row
    tr = ws_sal.max_row + 1
    ws_sal.cell(tr, 1, "TOTAL").font = BOLD
    ws_sal[f"K{tr}"] = f"=SUM(K2:K{tr-1})"
    
    _style_sheet(ws_sal)
    
    # --- Employee DB Sheet ---
    ws_db = wb.create_sheet("Employee DB")
    headers_db = ["Name", "Daily Rate", "Type"]
    ws_db.append(headers_db)
    
    for name_lower, info in sorted(employee_db.items(), key=lambda x: x[0]):
        ws_db.append([
            " ".join(w.capitalize() for w in name_lower.split()),
            info.get("daily_rate", 0),
            info.get("type", "weekly")
        ])
    
    _style_sheet(ws_db)
    
    # Save
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    wb.close()
    return output_path


def load_employee_db(prev_output_path):
    """Load employee DB and advance/arrears from previous output file if exists.
    
    Returns: (employee_db dict, carry_over dict {name_lower: {advance, arrears}})
    """
    employee_db = {}
    carry_over = {}
    
    if not prev_output_path or not os.path.exists(prev_output_path):
        return employee_db, carry_over
    
    try:
        wb = load_workbook(prev_output_path, data_only=True)
        
        # Load Employee DB tab
        if "Employee DB" in wb.sheetnames:
            ws = wb["Employee DB"]
            for r in range(2, ws.max_row + 1):
                name = str(ws.cell(r, 1).value or "").strip().lower()
                rate = ws.cell(r, 2).value or 0
                etype = str(ws.cell(r, 3).value or "weekly").strip().lower()
                if name:
                    employee_db[name] = {"daily_rate": int(rate), "type": etype}
        
        # Load Salary tab for advance/arrears carry-over
        if "Salary" in wb.sheetnames:
            ws = wb["Salary"]
            for r in range(2, ws.max_row + 1):
                name = str(ws.cell(r, 1).value or "").strip().lower()
                if name and name != "total":
                    adv = ws.cell(r, 9).value or 0
                    arr = ws.cell(r, 10).value or 0
                    carry_over[name] = {"advance": adv, "arrears": arr}
        
        wb.close()
    except Exception:
        pass
    
    return employee_db, carry_over


def _style_sheet(ws):
    """Apply borders, bold headers, and autofit columns."""
    for cell in ws[1]:
        cell.font = BOLD
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = CENTER
    
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 25)
